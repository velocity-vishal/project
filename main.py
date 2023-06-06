
import logging
import openpyxl
from fastapi import FastAPI
from constant import *
from models.ResponseModel import *
from http.client import INTERNAL_SERVER_ERROR, HTTPException
from fastapi.middleware.cors import CORSMiddleware

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*", "x_source"]
)


@app.get('/')
async def root():
    return {"message": "Hello World!!"}

@app.get('/convertExcelToSqlScript', response_model=ResponseModel)
async def convert_excel_to_sql_script():
    try:
        df = openpyxl.load_workbook(FILE_DEST)
        sheet = df.active
    except Exception as e:
        logger.error(e)
        raise HTTPException(status_code=404, detail=FILE_NOT_FOUND)
    
    data = []
    column_name = []
    query = f"INSERT INTO {DB_SCHEMA}.{TABLE_NAME} ("
    
    
    for col in sheet.iter_cols(1 , sheet.max_column):
        column_name.append(col[0].value)

    # print(column_name)
    for name in column_name:
        query += f"{name}, "
    query = query[:-2]
    query += ") VALUES ("

    info = []
    for row in range(1 , sheet.max_row):
        for col in sheet.iter_cols(1 , sheet.max_column):
            info.append(col[row].value)
            query += f"'{col[row].value}', "
        
        query = query[:-2]
        query += "), ("

    query = query[:-3]
    query += ";"

    return {"message": "success", "status": 200 , "data": query}



